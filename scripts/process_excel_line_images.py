from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
from pathlib import Path

import openpyxl
from PIL import Image


PROJECT_DIR = Path(r"C:\Users\kurtw\OneDrive\Documentos\New project\catalogo-produtos")
SEED_FILE = PROJECT_DIR / "imports" / "syt-05-2026" / "syt_seed_data.js"
IMPORT_ROOT = PROJECT_DIR / "imports" / "syt-05-2026"
SOURCE_DIR = Path(r"C:\Users\kurtw\OneDrive\Área de Trabalho\arquivos\price list syt 05.2026")


def safe_code(value: object) -> str:
    text = str(value or "").strip().upper()
    text = re.sub(r"[^A-Z0-9_-]+", "-", text)
    return text.strip("-")


def slugify(value: str) -> str:
    text = value.strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def image_extension(fmt: str | None) -> str:
    fmt = (fmt or "png").lower().replace("jpeg", "jpg")
    return fmt if fmt in {"jpg", "png", "gif", "bmp"} else "png"


def find_code_column(ws, header_name: str | None) -> int:
    expected = (header_name or "model").strip().lower()
    aliases = {expected, "model", "code", "item", "item no.", "item no", "model no.", "model no"}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)):
        for cell in row:
            label = str(cell.value or "").strip().lower()
            if label in aliases:
                return cell.column
    raise RuntimeError("Nao encontrei coluna de codigo/modelo na planilha.")


def load_seed() -> dict:
    text = SEED_FILE.read_text(encoding="utf-8")
    match = re.search(r"window\.CATALOGO_SEED_DATA\s*=\s*(\{.*\})\s*;?\s*$", text, re.S)
    if not match:
        raise RuntimeError("Nao consegui ler window.CATALOGO_SEED_DATA.")
    return json.loads(match.group(1))


def save_seed(data: dict) -> None:
    SEED_FILE.write_text(
        "window.CATALOGO_SEED_DATA = "
        + json.dumps(data, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )


def dimensions(path: Path) -> tuple[int, int]:
    with Image.open(path) as image:
        return image.size


def image_area(path: Path) -> int:
    width, height = dimensions(path)
    return width * height


def best_candidate(paths: list[Path]) -> Path:
    return max(paths, key=image_area)


def rel_to_path(url: str) -> Path:
    clean = url.replace("\\", "/")
    if clean.startswith("./"):
        clean = clean[2:]
    return PROJECT_DIR / clean


def rel_from_project(path: Path) -> str:
    return "./" + path.relative_to(PROJECT_DIR).as_posix()


def product_line(product: dict, catalogs_by_id: dict[str, dict]) -> str:
    return catalogs_by_id.get(product.get("source_file_id"), {}).get("line_name", "")


def extract_images(source_file: Path, output_dir: Path, code_header: str | None) -> list[dict]:
    output_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(source_file)
    ws = wb.active
    code_col = find_code_column(ws, code_header)
    rows: list[dict] = []
    seen: dict[str, int] = {}

    for idx, image in enumerate(getattr(ws, "_images", []), start=1):
        anchor = image.anchor._from
        row_number = anchor.row + 1
        code = safe_code(ws.cell(row_number, code_col).value)
        if not code:
            rows.append(
                {
                    "status": "SEM_CODIGO",
                    "row": row_number,
                    "code": "",
                    "file": "",
                    "image_index": idx,
                    "anchor_col": anchor.col + 1,
                }
            )
            continue

        seen[code] = seen.get(code, 0) + 1
        suffix = "" if seen[code] == 1 else f"-{seen[code]}"
        ext = image_extension(getattr(image, "format", None))
        file_name = f"{code}{suffix}.{ext}"
        (output_dir / file_name).write_bytes(image._data())
        rows.append(
            {
                "status": "OK",
                "row": row_number,
                "code": code,
                "file": file_name,
                "image_index": idx,
                "anchor_col": anchor.col + 1,
            }
        )

    with (output_dir / "_manifest.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["status", "row", "code", "file", "image_index", "anchor_col"],
        )
        writer.writeheader()
        writer.writerows(rows)

    return rows


def apply_if_larger(line_name: str, slug: str, extracted_dir: Path) -> tuple[int, int, Path]:
    data = load_seed()
    catalogs_by_id = {catalog["id"]: catalog for catalog in data.get("catalogs", [])}
    target_dir = IMPORT_ROOT / "product-images" / f"{slug}-excel"
    target_dir.mkdir(parents=True, exist_ok=True)
    report_file = extracted_dir / "_replace_if_larger_report.csv"

    rows: list[dict] = []
    changed = 0

    for product in data.get("products", []):
        code = safe_code(product.get("code"))
        if product_line(product, catalogs_by_id) != line_name:
            continue

        candidates = sorted(extracted_dir.glob(f"{code}.*")) + sorted(extracted_dir.glob(f"{code}-*.*"))
        if not candidates:
            rows.append({"code": code, "status": "SEM_IMAGEM_NOVA"})
            continue

        new_path = best_candidate(candidates)
        old_url = product.get("image_url") or ""
        if not old_url:
            rows.append({"code": code, "status": "SEM_IMAGEM_ANTIGA", "new_file": new_path.name})
            continue

        old_path = rel_to_path(old_url)
        if not old_path.exists():
            rows.append({"code": code, "status": "IMAGEM_ANTIGA_NAO_EXISTE", "old_url": old_url, "new_file": new_path.name})
            continue

        old_w, old_h = dimensions(old_path)
        new_w, new_h = dimensions(new_path)
        if new_w * new_h <= old_w * old_h:
            rows.append(
                {
                    "code": code,
                    "status": "MANTIDA_ANTIGA",
                    "old_url": old_url,
                    "old_size": f"{old_w}x{old_h}",
                    "new_file": new_path.name,
                    "new_size": f"{new_w}x{new_h}",
                }
            )
            continue

        target_path = target_dir / new_path.name
        shutil.copy2(new_path, target_path)
        new_url = rel_from_project(target_path)
        product["image_url"] = new_url
        product["gallery"] = [new_url]
        changed += 1
        rows.append(
            {
                "code": code,
                "status": "SUBSTITUIDA",
                "old_url": old_url,
                "old_size": f"{old_w}x{old_h}",
                "new_url": new_url,
                "new_size": f"{new_w}x{new_h}",
            }
        )

    with report_file.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["code", "status", "old_url", "old_size", "new_file", "new_url", "new_size"],
            extrasaction="ignore",
        )
        writer.writeheader()
        writer.writerows(rows)

    if changed:
        save_seed(data)
    return len(rows), changed, report_file


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True, help="Nome do arquivo .xlsx dentro da pasta price list.")
    parser.add_argument("--line", required=True, help="Nome exato da linha no catalogo.")
    parser.add_argument("--slug", help="Slug para pasta de saida. Padrao: linha normalizada.")
    parser.add_argument("--code-header", help="Cabecalho da coluna de codigo, se nao for Model.")
    args = parser.parse_args()

    source_file = SOURCE_DIR / args.file
    slug = args.slug or slugify(args.line)
    extracted_dir = IMPORT_ROOT / "excel-images" / slug

    extracted = extract_images(source_file, extracted_dir, args.code_header)
    evaluated, changed, report_file = apply_if_larger(args.line, slug, extracted_dir)

    print(f"Planilha: {source_file.name}")
    print(f"Linha do catalogo: {args.line}")
    print(f"Imagens extraidas: {sum(1 for row in extracted if row['status'] == 'OK')}")
    print(f"Produtos avaliados: {evaluated}")
    print(f"Imagens substituidas: {changed}")
    print(f"Pasta extraida: {extracted_dir}")
    print(f"Relatorio: {report_file}")


if __name__ == "__main__":
    main()
