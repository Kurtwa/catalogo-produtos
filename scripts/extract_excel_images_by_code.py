from __future__ import annotations

import csv
import re
from pathlib import Path

import openpyxl


SOURCE_FILE = Path(
    r"C:\Users\kurtw\OneDrive\Área de Trabalho\arquivos\price list syt 05.2026\A7 SERIES PRICE LIST.xlsx"
)
OUTPUT_DIR = Path(
    r"C:\Users\kurtw\OneDrive\Documentos\New project\catalogo-produtos\imports\syt-05-2026\excel-images\a7"
)


def safe_code(value: object) -> str:
    text = str(value or "").strip().upper()
    text = re.sub(r"[^A-Z0-9_-]+", "-", text)
    return text.strip("-")


def image_extension(fmt: str | None) -> str:
    fmt = (fmt or "png").lower().replace("jpeg", "jpg")
    if fmt not in {"jpg", "png", "gif", "bmp"}:
        return "png"
    return fmt


def find_model_column(ws) -> int:
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if str(cell.value or "").strip().lower() == "model":
                return cell.column
    raise RuntimeError("Nao encontrei a coluna Model na planilha.")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(SOURCE_FILE)
    ws = wb.active
    model_col = find_model_column(ws)

    rows = []
    seen: dict[str, int] = {}

    for idx, image in enumerate(ws._images, start=1):
        anchor = image.anchor._from
        row_number = anchor.row + 1
        raw_code = ws.cell(row_number, model_col).value
        code = safe_code(raw_code)

        if not code:
            rows.append(
                {
                    "status": "SEM_CODIGO",
                    "row": row_number,
                    "code": "",
                    "file": "",
                    "image_index": idx,
                    "anchor_col": anchor.col + 1,
                }
            )
            continue

        seen[code] = seen.get(code, 0) + 1
        suffix = "" if seen[code] == 1 else f"-{seen[code]}"
        ext = image_extension(getattr(image, "format", None))
        file_name = f"{code}{suffix}.{ext}"
        target = OUTPUT_DIR / file_name
        target.write_bytes(image._data())

        rows.append(
            {
                "status": "OK",
                "row": row_number,
                "code": code,
                "file": file_name,
                "image_index": idx,
                "anchor_col": anchor.col + 1,
            }
        )

    manifest = OUTPUT_DIR / "_manifest.csv"
    with manifest.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["status", "row", "code", "file", "image_index", "anchor_col"],
        )
        writer.writeheader()
        writer.writerows(rows)

    ok_count = sum(1 for row in rows if row["status"] == "OK")
    print(f"Planilha: {SOURCE_FILE.name}")
    print(f"Imagens extraidas: {ok_count}")
    print(f"Pasta: {OUTPUT_DIR}")
    print(f"Manifesto: {manifest}")


if __name__ == "__main__":
    main()
